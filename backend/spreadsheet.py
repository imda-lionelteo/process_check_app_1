import re
from tempfile import NamedTemporaryFile
from typing import Any, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from pandas import DataFrame, ExcelFile, Series
from pandas._libs.parsers import STR_NA_VALUES
from streamlit.logger import get_logger

logger = get_logger(__name__)

# Define constant for AI type
SELECTED_AI_TYPE: str = "Generative AI"

# Sheets to ignore when processing Excel data
IGNORED_SHEETS = ["Get Started", "Glossary", "Track your Progress"]

# Column indices for merged cell values - these are columns that may contain merged cells
# across multiple rows in the Excel spreadsheet
MERGED_CELL_COLUMNS = {
    "outcome_id": 0,
    "type_of_ai": 1,
    "outcomes": 2,
    "process_to_achieve_outcomes": 4,
    "nature_of_evidence": 5,
    "evidence": 6,
}

# Define column indices for process check data - these map to specific columns
# in the Excel spreadsheet containing process check information
PROCESS_CHECK_COLUMNS = {
    "outcome_id": 0,
    "type_of_ai": 1,
    "outcomes": 2,
    "process_id": 3,
    "process_to_achieve_outcomes": 4,
    "nature_of_evidence": 5,
    "evidence": 6,
    "implementation": 7,
    "elaboration": 8,
}


# ------------------------------------------------------------------------------
# Read excel file for principles data - using pandas
# ------------------------------------------------------------------------------
def extract_principle_description(df: DataFrame) -> str:
    """
    Extract the principle description from the DataFrame.

    Args:
        df: DataFrame containing the principle data

    Returns:
        The principle description as a string
    """
    return str(df.iloc[1, 0])


def is_valid_process_id(process_id: Any) -> bool:
    """
    Check if the process ID has a valid format.

    Args:
        process_id: The process ID to validate

    Returns:
        True if the process ID is valid, False otherwise
    """
    return (
        isinstance(process_id, (str, int, float))
        and re.match(r"^\d+\.\d+\.\d+$", str(process_id)) is not None
    )


def matches_ai_type_filter(type_of_ai: Any, ai_type_filter: str) -> bool:
    """
    Check if the AI type matches the filter.

    Args:
        type_of_ai: The AI type from the row
        ai_type_filter: The AI type to filter by

    Returns:
        True if the AI type matches the filter, False otherwise
    """
    return isinstance(type_of_ai, str) and ai_type_filter in type_of_ai


def parse_process_check_row(
    row: Series, merged_cell_values: dict[str, Any], ai_type_filter: str
) -> Optional[dict[str, Any]]:
    """
    Parse a row from the process check spreadsheet, handling merged cells.

    Args:
        row: A DataFrame row containing process check data
        merged_cell_values: Values from merged cells carried over from previous rows
        ai_type_filter: The AI type to filter questions by

    Returns:
        Process check data from the row or None if invalid
    """
    # Dynamically get all merged cell values
    merged_values = {key: merged_cell_values[key] for key in MERGED_CELL_COLUMNS}

    process_id = row.iloc[PROCESS_CHECK_COLUMNS["process_id"]]

    # Validate process ID format
    if not is_valid_process_id(process_id):
        return None

    # Filter by AI type
    if not matches_ai_type_filter(merged_values["type_of_ai"], ai_type_filter):
        return None

    return {
        **merged_values,
        "implementation": (
            row.iloc[PROCESS_CHECK_COLUMNS["implementation"]]
            if pd.notna(row.iloc[PROCESS_CHECK_COLUMNS["implementation"]])
            else None
        ),
        "elaboration": (
            row.iloc[PROCESS_CHECK_COLUMNS["elaboration"]]
            if pd.notna(row.iloc[PROCESS_CHECK_COLUMNS["elaboration"]])
            else ""
        ),
    }


def process_excel_principles_data(excel_data: ExcelFile) -> dict[str, dict[str, Any]]:
    """
    Process the Excel file and extract principle data with process checks.

    This function reads an Excel file containing AI principles and process checks data.
    It processes each sheet (except those in IGNORED_SHEETS) to extract principle descriptions and
    their associated process checks. The data is organized into a nested dictionary structure.

    Args:
        excel_data: The Excel file data containing principles and process checks

    Returns:
        A dictionary where:
            - Keys are sheet names (principle IDs)
            - Values are dictionaries containing:
                - 'principle_description': Description of the principle
                - 'process_checks': Dictionary of process checks for that principle

    Example return format:
        {
            'P1': {
                'principle_description': 'Description of Principle 1',
                'process_checks': {
                    'P1.1': {process check data},
                    'P1.2': {process check data}
                }
            }
        }
    """
    principles_data = {}

    for sheet_name in excel_data.sheet_names:
        if str(sheet_name) in IGNORED_SHEETS:
            continue

        # Remove "N/A" from the default NA values
        accepted_na_values = STR_NA_VALUES - {"N/A"}
        df = pd.read_excel(
            excel_data,
            sheet_name=sheet_name,
            keep_default_na=False,
            na_values=list(accepted_na_values),
        )
        principle_data = process_single_principle_sheet(df, sheet_name)

        if principle_data:
            principles_data[str(sheet_name)] = principle_data

    logger.info(f"Extracted data for {len(principles_data)} principles")
    return principles_data


def process_single_principle_sheet(
    df: DataFrame, sheet_name: Any
) -> Optional[dict[str, Any]]:
    """
    Process a single principle sheet from the Excel file.

    Args:
        df: DataFrame containing the principle data
        sheet_name: Name of the sheet being processed

    Returns:
        Dictionary with principle description and process checks, or None if error
    """
    try:
        principle_description = extract_principle_description(df)
        process_checks = {}
        merged_cell_values = {key: None for key in MERGED_CELL_COLUMNS}

        for index, row in df.iterrows():
            if isinstance(index, int) and index < 1:  # Skip header row
                continue

            # Update merged cell values
            merged_cell_values = update_merged_cell_values(row, merged_cell_values)

            # Parse the row data
            process_check = parse_process_check_row(
                row, merged_cell_values, SELECTED_AI_TYPE
            )

            # Add valid process checks to the collection
            if process_check and process_check["process_to_achieve_outcomes"]:
                process_id = str(row.iloc[PROCESS_CHECK_COLUMNS["process_id"]])
                process_checks[process_id] = process_check

        # Return principle data if it has process checks
        if process_checks:
            return {
                "principle_description": principle_description,
                "process_checks": process_checks,
            }
        return None

    except Exception as e:
        logger.error(f"Error processing principle sheet {sheet_name}: {str(e)}")
        return None


def read_principles_from_excel(excel_file) -> dict[str, dict[str, Any]]:
    """
    Load AI principles data from an Excel file.

    Args:
        excel_file): The Excel file containing principles data.
            Can be a file path string, bytes object, or file-like object.

    Returns:
        dict[str, dict[str, Any]]: A nested dictionary containing the principles data.
            The outer dictionary is keyed by principle IDs (str), and each inner dictionary
            contains the principle's details including:
            - principle_description (str): Description of the principle
            - process_checks (dict): Dictionary of process checks keyed by process ID

    Raises:
        Exception: If there is an error reading or processing the Excel file.
            The error will be logged and an empty dict will be returned.
    """
    try:
        return process_excel_principles_data(pd.ExcelFile(excel_file))
    except Exception as e:
        logger.error(f"Error reading principles from Excel file: {str(e)}")
        return {}


def update_merged_cell_values(
    row: Series, merged_cell_values: dict[str, Any]
) -> dict[str, Any]:
    """
    Update the merged cell values dictionary with non-NA values from the current row.

    Args:
        row: A DataFrame row containing process check data
        merged_cell_values: Values from merged cells carried over from previous rows

    Returns:
        Updated merged cell values dictionary
    """
    updated_values = merged_cell_values.copy()

    for key, col_idx in MERGED_CELL_COLUMNS.items():
        if pd.notna(row.iloc[col_idx]):
            updated_values[key] = row.iloc[col_idx]

    return updated_values


# ------------------------------------------------------------------------------
# Export excel file - using openpyxl
# ------------------------------------------------------------------------------
def carry_forward_merged_values(row, last_values: dict, column_indices: dict) -> dict:
    """
    Carry forward values from merged cells in Excel.

    Args:
        row: The current row being processed
        last_values: Dictionary of the last non-None values seen
        column_indices: Dictionary mapping column names to indices

    Returns:
        Dictionary with current values for this row
    """
    current_values = last_values.copy()

    # Helper function to safely get cell value
    def get_cell_value(idx):
        return row[idx].value if idx < len(row) else None

    # Update each tracked value if present in current row
    for field, idx in column_indices.items():
        if field in current_values:
            val = get_cell_value(idx)
            if val is not None:
                current_values[field] = (
                    str(val) if field in ["outcome_id", "process_id"] else val
                )

    return current_values


def export_excel(template_path: str, updates: Optional[dict] = None) -> bytes:
    """
    Creates an Excel file with updated values based on a template.

    Args:
        template_path: Path to the Excel template file
        updates: Nested dictionary with structure {outcome_id: {process_id: {fields...}}}

    Returns:
        The Excel file content as bytes
    """
    workbook = load_workbook(template_path)

    # Set active page to the second worksheet (index 1)
    workbook.active = 2

    if updates:
        updates = stringify_dict_keys(updates)
        update_process_check_rows(workbook, updates)

    with NamedTemporaryFile() as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        return tmp.read()


def get_process_data(current_values: dict, updates: dict) -> dict | None:
    """
    Retrieve process data from the updates dictionary using outcome_id and process_id as keys.

    Args:
        current_values: Dictionary containing current row values including outcome_id and process_id
        updates: Nested dictionary with structure {outcome_id: {process_id: {field1: value1, ...}}}

    Returns:
        dict | None: The process data dictionary if found, or None if either outcome_id or process_id
        is missing or not found in the updates dictionary
    """
    outcome_id = current_values.get("outcome_id")
    process_id = current_values.get("process_id")

    if not outcome_id or not process_id:
        return None

    return updates.get(outcome_id, {}).get(process_id)


def is_matching_process(process: dict, sheet_name: str, current_values: dict) -> bool:
    """
    Check if the process data matches the current row.

    Args:
        process: Process data dictionary
        sheet_name: Name of the current sheet
        current_values: Dictionary containing current row values

    Returns:
        True if all fields match, False otherwise
    """
    return (
        process.get("principle_key") == sheet_name
        and process.get("outcomes") == current_values.get("outcomes")
        and process.get("process_to_achieve_outcomes")
        == current_values.get("process_to_achieve_outcomes")
        and process.get("nature_of_evidence")
        == current_values.get("nature_of_evidence")
        and process.get("evidence") == current_values.get("evidence")
    )


def process_sheet_rows(
    sheet, sheet_name: str, updates: dict, column_indices: dict
) -> None:
    """
    Process rows in a worksheet to update implementation and elaboration fields.

    Args:
        sheet: The worksheet being processed
        sheet_name: Name of the current sheet
        updates: Dictionary containing update data
        column_indices: Dictionary mapping column names to indices
    """
    # Track last non-None values for merged cells
    last_values = {}
    for field in column_indices.keys():
        last_values[field] = None

    for row in sheet.iter_rows(min_row=2):
        # Skip empty rows
        if all(cell.value is None for cell in row):
            continue

        # Update and carry forward merged cell values
        current_values = carry_forward_merged_values(row, last_values, column_indices)
        last_values = current_values.copy()

        # Get the process data for this row if it exists
        process = get_process_data(current_values, updates)
        if not process:
            continue

        # Update the row if all fields match
        if is_matching_process(process, sheet_name, current_values):
            update_row_values(row, process, column_indices)


def stringify_dict_keys(nested_dict: dict) -> dict:
    """
    Convert all dictionary keys to strings in a nested dictionary.

    Args:
        nested_dict: A nested dictionary with potentially non-string keys

    Returns:
        A nested dictionary with all string keys
    """
    if not nested_dict:
        return {}
    return {
        str(k): {str(pk): v for pk, v in vdict.items()}
        for k, vdict in nested_dict.items()
    }


def update_process_check_rows(workbook: Workbook, updates: dict) -> None:
    """
    Updates elaboration and implementation fields in the Excel workbook.

    Args:
        workbook: An openpyxl workbook object
        updates: Nested dictionary with structure {outcome_id: {process_id: {fields...}}}
    """
    if not updates:
        return

    for sheet_name in workbook.sheetnames:
        if sheet_name in IGNORED_SHEETS:
            continue
        sheet = workbook[sheet_name]

        # Process each sheet, starting from row 2 (skipping header)
        process_sheet_rows(sheet, sheet_name, updates, PROCESS_CHECK_COLUMNS)


def update_row_values(row, process: dict, column_indices: dict) -> None:
    """
    Update implementation and elaboration values in the row.

    Args:
        row: The row being updated
        process: Process data dictionary
        column_indices: Dictionary mapping column names to indices
    """
    if "elaboration" in process:
        row[column_indices["elaboration"]].value = process["elaboration"]
    if "implementation" in process:
        row[column_indices["implementation"]].value = process["implementation"]
